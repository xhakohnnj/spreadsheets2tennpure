#
# XLSXからリリースデータのリストに変換
#
from itertools import count
import openpyxl
from .lib import ReleaseDataListCreaterFromCsv


# ファイルから作成
def FromFile( file_path, sheet_name, date_start_str, date_end_str ):
    wb = openpyxl.load_workbook( file_path )
    if wb is None:
        return None

    sheet = wb[sheet_name]
    if sheet is None:
        return None

    row_base = 3
    row_last = 0
    col_base = 2
    col_last = 5

    # rowの最後の位置を検索
    for cnt in count():
        # 日付がなければ終了
        cell = sheet.cell( row=row_base+cnt, column=col_base )
        if cell.value is None:
            break
        row_last = row_base+cnt

    datas = []
    if 0 < row_last:
        # セルの値を取得
        cells = sheet.iter_rows( min_row=row_base, max_row=row_last, min_col=col_base, max_col=col_last )

        # 値だけの配列に変換=データ
        def get_value_list(t_2d):
            return ([[cell.value for cell in row] for row in t_2d])
        datas = get_value_list( cells )

    # ファイルはいらなくなったのでクローズ
    wb.close()

    # データを作成して出力
    return ReleaseDataListCreaterFromCsv.Create( datas, date_start_str, date_end_str ) if datas != None and 0 < len(datas) else None
