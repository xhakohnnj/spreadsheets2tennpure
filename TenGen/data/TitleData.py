#
# タイトルデータ
#
from enum import IntEnum, auto
from itertools import count
from datetime import datetime
import openpyxl
from ..util import Util
from .. import Error
from .. import ErrorLog


# データ本体
class Body:
    def __init__( self, date:datetime=None, time:datetime=None, titleName:str=None, supplement:str=None, userData1:str=None, userData2:str=None, userData3:str=None ):
      self.date = date              # リリース日
      self.time = time              # リリース時間
      self.titleName = titleName    # タイトル名
      self.supplement = supplement  # 補足
      self.userData1 = userData1    # なんでもデータ 1
      self.userData2 = userData2    # なんでもデータ 2
      self.userData3 = userData3    # なんでもデータ 3

# データ本体のリストをXlsxから作成
def BodyListFromXlsx( file_path:str, sheet_name:str, date_start_str:str, date_end_str:str ):
    wb = openpyxl.load_workbook( file_path )
    if wb is None:
        return None

    sheet = wb[sheet_name]
    if sheet is None:
        return None

    row_base = 4
    row_last = 0
    col_base = 3
    col_last = 9 # ここなんかうまいことできたらいいんだけど.

    # rowの最後の位置を検索
    for cnt in count():
        # 日付がなければ終了
        cell = sheet.cell( row=row_base+cnt, column=col_base )
        if cell.value is None:
            break
        row_last = row_base+cnt

    csv = [] # 正確にはCSVではないんだけど.
    if 0 < row_last:
        # セルの値を取得
        cells = sheet.iter_rows( min_row=row_base, max_row=row_last, min_col=col_base, max_col=col_last )

        # 値だけの配列に変換=データ
        def get_value_list(t_2d):
            return ([[cell.value for cell in row] for row in t_2d])
        csv = get_value_list( cells )

    # ファイルはいらなくなったのでクローズ
    wb.close()

    # データを作成
    return BodyListFromCsv( sheet_name, csv, date_start_str, date_end_str ) if csv != None and 0 < len(csv) else None

def BodyListFromCsv( sheet_name:str, items:list, date_start_str:str, date_end_str:str ):
    date_start = Util.StrToDate( date_start_str )
    date_end = Util.StrToDate( date_end_str )

    title_data_list = []
    cnt = 0
    for item in items:
        error_logs = []
        title_data = FromCSV( sheet_name, cnt, item, error_logs )

        if( 0 < len(error_logs) ):
            for error_log in error_logs:
                ErrorLog.Logs.Add( error_log )
        else:
            # 日付の範囲
            if date_start <= title_data.date <= date_end:
                title_data_list.append( title_data )
        cnt = cnt + 1

    # 日付順でソート.
    def ToSort( r_data ):
        return r_data.date
    title_data_list.sort( key=ToSort )

    return title_data_list

def FromCSV( sheet_name:str, cnt:int, csv_data:list, error_logs:list ):
    class eCSV(IntEnum):
        ReleaseDate = 0         # リリース日
        ReleaseTime = auto()    # リリース時間
        TitleName   = auto()    # タイトル名
        Supplement  = auto()    # 補足
        UserData1   = auto()    # なんでもデータ.名前が思いつかず適当.
        UserData2   = auto()    # ...
        UserData3   = auto()    # ...
        Num         = auto()

    if( len(csv_data) != eCSV.Num ):
      return None

    # エラーチェック
    def ErrorCheck( err_condition:bool, error_id:Error.eID, msg:str ):
        if err_condition:
            error_logs.append( ErrorLog.Log(error_id,msg) )
    def GetExtendInfo(end_newline_del=False):
        ret = ""
        if csv_data[eCSV.TitleName] is not None: ret += f"タイトル名: {csv_data[eCSV.TitleName]}\n"
        if end_newline_del: ret = ret.rstrip('\n')
        return ret
    #if title_data is None: @todo なんかエラー考えなければ.
    ErrorCheck( csv_data[eCSV.TitleName] is None, Error.eID.ExportCommon_TitleIsNone, f"「{sheet_name}」シート\n{cnt+1} 番目" )
    ErrorCheck( csv_data[eCSV.ReleaseDate] is None, Error.eID.ExportCommon_DateIsNone, f"「{sheet_name}」シート\n{cnt+1} 番目\n{GetExtendInfo(True)}" )
    ErrorCheck( type(csv_data[eCSV.ReleaseDate]) is not str, Error.eID.ExportCommon_DateIsNotStrings, f"「{sheet_name}」シート\n{cnt+1} 番目\n{GetExtendInfo(True)}" )
    if csv_data[eCSV.ReleaseTime]:
        ErrorCheck( type(csv_data[eCSV.ReleaseTime]) is not str, Error.eID.ExportCommon_TimeIsNotStrings, f"「{sheet_name}」シート\n{cnt+1} 番目\n{GetExtendInfo(True)}" )

    if( 0 < len(error_logs) ):
        return None

    body = Body()
    body.date       = datetime.strptime( csv_data[eCSV.ReleaseDate], '%Y/%m/%d' )
    body.time       = datetime.strptime( csv_data[eCSV.ReleaseTime], '%H:%M' ) if csv_data[eCSV.ReleaseTime] is not None else None
    body.titleName  = csv_data[eCSV.TitleName]
    body.supplement = csv_data[eCSV.Supplement]
    body.userData1  = csv_data[eCSV.UserData1]
    body.userData2  = csv_data[eCSV.UserData2]
    body.userData3  = csv_data[eCSV.UserData3]
    return body
